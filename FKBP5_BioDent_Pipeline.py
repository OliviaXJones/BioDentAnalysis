import re
import json
import sys
import pandas as pd
import openpyxl
from pathlib import Path
import shutil

from PyQt5.QtWidgets import (
    QApplication, QDialog, QVBoxLayout, QHBoxLayout, QFormLayout,
    QLabel, QLineEdit, QPushButton, QFileDialog, QDialogButtonBox,
    QMessageBox, QTableWidget, QTableWidgetItem, QHeaderView,
    QInputDialog, QWidget, QFrame, QSizePolicy
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont

# --- 1. SETTINGS ---

COLUMNS_TO_AVERAGE = [
    "1st Cycle Touchdown Distance (TDD 1st) - um",
    "Avg Loading Slope (Avg LS 1st-L) - N/um",
    "Avg Unloading Slope (Avg US 1st-L) - N/um",
    "Avg Energy Dissipated (Avg ED 3rd-L) - uJ",
    "Avg Creep Indentation Distance (Avg CID 1st-L) - um",
    "Indentation Distance Increase (IDI 1st-L) - um",
    "Total Indentation Distance (TID 1st-L) - um",
    "1st Cycle Creep Indentation Distance (CID 1st) - um",
    "1st Cycle Unloading Slope (US 1st) - N/um",
    "1st Cycle Indentation Distance (ID 1st) - um"
]

CONFIG_PATH = Path(__file__).parent / "study_config.json"


# --- 2. CONFIG DIALOG ---

class ConfigDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("FKBP5 BioDent Analysis Pipeline")
        self.setMinimumWidth(640)
        self.setWindowModality(Qt.ApplicationModal)
        self._config = {}

        existing = {}
        if CONFIG_PATH.exists():
            try:
                with open(CONFIG_PATH) as f:
                    existing = json.load(f)
            except Exception:
                pass

        root_layout = QVBoxLayout(self)
        root_layout.setSpacing(16)
        root_layout.setContentsMargins(28, 24, 28, 24)

        # Title
        title = QLabel("Study Configuration")
        title.setFont(QFont("Arial", 15, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        root_layout.addWidget(title)

        divider = QFrame()
        divider.setFrameShape(QFrame.HLine)
        divider.setFrameShadow(QFrame.Sunken)
        root_layout.addWidget(divider)

        # Form
        form = QFormLayout()
        form.setSpacing(12)
        form.setLabelAlignment(Qt.AlignRight | Qt.AlignVCenter)

        self._study_name = QLineEdit(existing.get("study_name", ""))
        self._study_name.setPlaceholderText("e.g. FKBP5_Cohort1")
        form.addRow("Study Name:", self._study_name)

        self._raw_data, raw_row = self._path_row(
            existing.get("raw_data_root", ""), folder=True)
        form.addRow("Raw Data Folder:", raw_row)

        self._master, master_row = self._path_row(
            existing.get("master_file", ""), folder=False)
        form.addRow("Master File (.xlsx):", master_row)

        self._output, output_row = self._path_row(
            existing.get("output_folder", ""), folder=True)
        form.addRow("Output Folder:", output_row)

        root_layout.addLayout(form)

        # Buttons
        root_layout.addSpacing(8)
        btn_box = QDialogButtonBox()
        run_btn = btn_box.addButton(
            "Run Pipeline", QDialogButtonBox.AcceptRole)
        run_btn.setFixedHeight(36)
        run_btn.setFont(QFont("Arial", 10, QFont.Bold))
        cancel_btn = btn_box.addButton("Cancel", QDialogButtonBox.RejectRole)
        cancel_btn.setFixedHeight(36)
        btn_box.accepted.connect(self._on_accept)
        btn_box.rejected.connect(self.reject)
        root_layout.addWidget(btn_box)

    def _path_row(self, value, folder):
        edit = QLineEdit(value)
        edit.setPlaceholderText("Click Browse or type a path...")
        edit.setMinimumWidth(360)

        btn = QPushButton("Browse...")
        btn.setFixedWidth(90)
        if folder:
            btn.clicked.connect(lambda: self._pick_folder(edit))
        else:
            btn.clicked.connect(lambda: self._pick_file(edit))

        row_widget = QWidget()
        h = QHBoxLayout(row_widget)
        h.setContentsMargins(0, 0, 0, 0)
        h.setSpacing(8)
        h.addWidget(edit)
        h.addWidget(btn)
        return edit, row_widget

    def _pick_folder(self, edit):
        start = edit.text() or str(Path.home())
        path = QFileDialog.getExistingDirectory(self, "Select Folder", start)
        if path:
            edit.setText(path)

    def _pick_file(self, edit):
        start = edit.text() or str(Path.home())
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Master Excel File", start,
            "Excel Files (*.xlsx *.xlsm)"
        )
        if path:
            edit.setText(path)

    def _on_accept(self):
        name = self._study_name.text().strip()
        raw = self._raw_data.text().strip()
        master = self._master.text().strip()
        output = self._output.text().strip()

        if not all([name, raw, master, output]):
            QMessageBox.warning(self, "Missing Fields",
                                "Please fill in all four fields before running.")
            return

        self._config = {
            "study_name":    name,
            "raw_data_root": raw,
            "master_file":   master,
            "output_folder": output,
        }
        with open(CONFIG_PATH, "w") as f:
            json.dump(self._config, f, indent=4)
        self.accept()

    def get_config(self):
        return self._config


# --- 3. MANUAL SELECTION DIALOG ---

class ManualSelectionDialog(QDialog):
    def __init__(self, mouse_id, group, target_col):
        super().__init__()
        self.setWindowTitle(f"Data Selection: {mouse_id}")
        self.setMinimumSize(960, 480)
        self.setWindowModality(Qt.ApplicationModal)
        self._group_index = list(group.index)
        self._selected = []

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 16, 20, 16)
        layout.setSpacing(12)

        label = QLabel(f"Select the 4 rows to keep for <b>{mouse_id}</b>:")
        label.setFont(QFont("Arial", 12))
        layout.addWidget(label)

        self._table = QTableWidget()
        self._table.setColumnCount(5)
        self._table.setHorizontalHeaderLabels(
            ["Index", "Meas #", "ID 1st (um)", "Notes", "Keep?"])
        self._table.setRowCount(len(group))
        self._table.setEditTriggers(QTableWidget.NoEditTriggers)
        self._table.setSelectionMode(QTableWidget.NoSelection)
        self._table.setAlternatingRowColors(True)
        self._table.verticalHeader().setVisible(False)
        self._table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self._table.setColumnWidth(0, 70)
        self._table.setColumnWidth(1, 80)
        self._table.setColumnWidth(2, 130)
        self._table.setColumnWidth(4, 80)

        for row_i, (idx, row) in enumerate(group.iterrows()):
            self._table.setItem(row_i, 0, QTableWidgetItem(str(idx)))
            self._table.setItem(row_i, 1, QTableWidgetItem(
                str(row["Measurement #"])))
            self._table.setItem(
                row_i, 2, QTableWidgetItem(str(row[target_col])))
            self._table.setItem(
                row_i, 3, QTableWidgetItem(str(row["Notes"])[:80]))

            keep = QTableWidgetItem()
            keep.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            keep.setCheckState(Qt.Unchecked)
            keep.setTextAlignment(Qt.AlignCenter)
            self._table.setItem(row_i, 4, keep)

        layout.addWidget(self._table)

        confirm_btn = QPushButton("Confirm Selection")
        confirm_btn.setFixedHeight(40)
        confirm_btn.setFont(QFont("Arial", 11, QFont.Bold))
        confirm_btn.setStyleSheet(
            "QPushButton { background-color: #4CAF50; color: white; border-radius: 5px; }"
            "QPushButton:hover { background-color: #43A047; }"
        )
        confirm_btn.clicked.connect(self._on_confirm)
        layout.addWidget(confirm_btn)

    def _on_confirm(self):
        self._selected = [
            self._group_index[r]
            for r in range(self._table.rowCount())
            if self._table.item(r, 4).checkState() == Qt.Checked
        ]
        self.accept()

    def get_selected_indices(self):
        return self._selected


# --- 4. HELPER DIALOGS ---

def ask_small_group_action(mouse_id, count):
    """Returns 'a' (average), 's' (skip), or 'r' (rename)."""
    msg = QMessageBox()
    msg.setWindowTitle("Insufficient Measurements")
    msg.setText(f"<b>{mouse_id}</b> only has {count} measurement(s).")
    msg.setInformativeText("How would you like to proceed?")
    avg_btn = msg.addButton("Average Anyway", QMessageBox.AcceptRole)
    skip_btn = msg.addButton("Skip",           QMessageBox.RejectRole)
    rename_btn = msg.addButton("Rename / Realign", QMessageBox.ActionRole)
    msg.setDefaultButton(avg_btn)
    msg.exec_()
    clicked = msg.clickedButton()
    if clicked == rename_btn:
        return 'r'
    if clicked == skip_btn:
        return 's'
    return 'a'


def ask_new_id(mouse_id, auto_detected=None):
    """Prompts for a corrected mouse ID. Returns the new ID string or None."""
    default = auto_detected or ""
    hint = (
        f"Auto-detected ID from notes: <b>{auto_detected}</b><br><br>" if auto_detected else "")
    text, ok = QInputDialog.getText(
        None,
        "Rename Mouse ID",
        f"{hint}Enter the correct ID for <b>{mouse_id}</b>:",
        text=default
    )
    if ok and text.strip():
        return text.strip().upper()
    return None


def ask_not_found_action(mouse_code):
    """
    Shown when a mouse ID is not found in any Excel sheet.
    Returns a new ID string to retry with, or None to skip.
    """
    msg = QMessageBox()
    msg.setWindowTitle("Mouse Not Found in Master")
    msg.setText(f"<b>'{mouse_code}'</b> was not found in any sheet.")
    msg.setInformativeText(
        "Would you like to rename and retry, or skip this mouse?")
    rename_btn = msg.addButton("Rename / Retry", QMessageBox.AcceptRole)
    skip_btn = msg.addButton("Skip",           QMessageBox.RejectRole)
    msg.setDefaultButton(skip_btn)
    msg.exec_()

    if msg.clickedButton() == rename_btn:
        text, ok = QInputDialog.getText(
            None,
            "Enter Mouse ID",
            "Enter the ID exactly as it appears in Excel:"
        )
        if ok and text.strip():
            return text.strip()
    return None


# --- 5. HELPER FUNCTIONS ---

def parse_squashed_code(code):
    """Parses IDs like 'W12M10', 'H24F5', or '2W10M10' for CSV grouping."""
    code = str(code).strip().upper()
    match = re.match(r"(\d+)?([A-Z])(\d+)([MF])(\d+)", code)
    if match:
        gen_num, gen_char, age, sex_char, mouse_num = match.groups()
        gen_map = {'W': 'Wildtype', 'M': 'Mutant', 'H': 'Heterozygous'}
        genotype = gen_map.get(gen_char, "Unknown")
        sex = "Male" if sex_char == 'M' else "Female"
        if gen_num:
            genotype = f"{gen_num}{genotype}"
        return genotype, age, sex, mouse_num
    return None, None, None, None


def get_genotype_prefix(code):
    if not code:
        return ""
    genotype_full, _, _, _ = parse_squashed_code(code)
    if genotype_full:
        return re.sub(r'\d+', '', genotype_full)
    return ""


def get_sex(code):
    return 'F' if 'F' in str(code).upper() else 'M'


# --- 6. DATA CLEANING & AVERAGING ---

def clean_and_average_data(txt_file_paths):
    all_data = []
    for path in txt_file_paths:
        try:
            temp_df = pd.read_csv(path, sep='\t', encoding='latin1')
            all_data.append(temp_df)
        except Exception as e:
            print(f"Skipping {path.name}: {e}")

    if not all_data:
        return pd.DataFrame()

    df = pd.concat(all_data, ignore_index=True)
    df.columns = [c.replace('µ', 'u') if isinstance(
        c, str) else c for c in df.columns]
    df = df.rename(columns={'Sample/Location': 'ID'})

    ID_REGEX = r"2?[WMH][.\-_]?\d+[.\-_]?[MF][.\-_]?\d+"

    def fix_mislabeled_id(row):
        notes = str(row['Notes']).strip().upper()
        match = re.search(ID_REGEX, notes)
        if match:
            true_id = re.sub(r"[.\-_]", "", match.group(0))
            if true_id != str(row['ID']).strip().upper():
                return true_id
        return row['ID']

    df['ID'] = df.apply(fix_mislabeled_id, axis=1)
    target_val_col = "1st Cycle Indentation Distance (ID 1st) - um"
    df = df.drop_duplicates(
        subset=['ID', 'Measurement #', target_val_col], keep='first')

    ignore_keywords = ["ignore", "do not use", "disregard", "don't", "ignor"]
    mask = df['Notes'].str.contains(
        '|'.join(ignore_keywords), case=False, na=False)
    df_cleaned = df[~mask].copy()
    df_cleaned = df_cleaned.dropna(subset=['ID'])
    df_cleaned['ID'] = df_cleaned['ID'].astype(str)

    final_averages = []
    unique_ids = sorted(df_cleaned['ID'].unique())
    i = 0
    while i < len(unique_ids):
        mouse_id = unique_ids[i]
        group = df_cleaned[df_cleaned['ID'] == mouse_id]
        mouse_id_str = str(mouse_id).strip()

        if len(group) > 4:
            dlg = ManualSelectionDialog(mouse_id_str, group, target_val_col)
            dlg.exec_()
            indices = dlg.get_selected_indices()
            if indices:
                group = group.loc[indices]

        elif 0 < len(group) < 4:
            action = ask_small_group_action(mouse_id_str, len(group))
            if action == 'r':
                auto_id = None
                for idx, row in group.iterrows():
                    scanned = fix_mislabeled_id(row)
                    if scanned != row['ID']:
                        auto_id = scanned
                        break
                new_id = ask_new_id(mouse_id_str, auto_id)
                if new_id:
                    new_id = re.sub(r"[.\-_]", "", new_id)
                    df_cleaned.loc[group.index, 'ID'] = new_id
                    unique_ids = sorted(df_cleaned['ID'].unique())
                    continue
            elif action == 's':
                i += 1
                continue

        for col in COLUMNS_TO_AVERAGE:
            group[col] = pd.to_numeric(group[col], errors='coerce')

        avg_values = group[COLUMNS_TO_AVERAGE].mean()
        avg_values['ID'] = mouse_id_str
        final_averages.append(avg_values)
        i += 1

    return pd.DataFrame(final_averages)


# --- 7. MASTER SYNC ---

def sync_to_master(df_source, master_path):
    wb = openpyxl.load_workbook(master_path)
    sync_count = 0
    synced_ids = set()
    all_processed_ids = set(df_source['ID'].astype(str).str.strip())

    all_sheets = [
        s for s in wb.sheetnames
        if s.lower() not in ["summary", "notes", "calculations"]
    ]

    for _, row_data in df_source.iterrows():
        mouse_code = str(row_data['ID']).strip()
        found_match = False

        while True:
            genotype = get_genotype_prefix(mouse_code)
            sex = get_sex(mouse_code)
            start_col = 1 if sex == 'M' else 13

            target_sheets = [
                s for s in all_sheets if genotype.lower() in s.lower()]
            other_sheets = [s for s in all_sheets if s not in target_sheets]
            search_order = target_sheets + other_sheets

            for sheet_name in search_order:
                ws = wb[sheet_name]
                for r in range(2, max(ws.max_row + 1, 50)):
                    cell_val = ws.cell(row=r, column=start_col).value
                    clean_cell = str(cell_val).strip(
                    ).upper() if cell_val else ""
                    if clean_cell == mouse_code.strip().upper():
                        for j, col_name in enumerate(COLUMNS_TO_AVERAGE):
                            ws.cell(row=r, column=start_col + 1 +
                                    j).value = row_data[col_name]
                        sync_count += 1
                        synced_ids.add(mouse_code)
                        found_match = True
                        break
                if found_match:
                    break

            if found_match:
                break

            new_name = ask_not_found_action(mouse_code)
            if new_name:
                mouse_code = new_name
            else:
                break

    wb.save(master_path)
    print(f"Synced {sync_count} mice to master.")

    unmatched_ids = all_processed_ids - synced_ids
    if unmatched_ids:
        print("The following mice were processed but not found in the master list:")
        for m_id in sorted(unmatched_ids):
            print(f"  - {m_id}")
    else:
        print("All processed mice were successfully matched in the master list.")


# --- 8. CSV GENERATION ---

def generate_grouped_tables(df, csv_output_dir):
    sort_priority = ['Wildtype', 'Mutant', 'Heterozygous']
    folder_genotype = csv_output_dir / "Analysis_By_Genotype"
    folder_lineage = csv_output_dir / "Analysis_By_Lineage"
    folder_genotype.mkdir(parents=True, exist_ok=True)
    folder_lineage.mkdir(parents=True, exist_ok=True)

    for metric in COLUMNS_TO_AVERAGE:
        clean_metric = re.sub(
            r'[\-\/\(\)]', '', metric).replace("  ", " ").strip()
        for age in df['Age_Extracted'].dropna().unique():
            for sex in ['Male', 'Female']:
                base_subset = df[
                    (df['Sex_Extracted'] == sex) & (df['Age_Extracted'] == age)
                ].copy()
                if base_subset.empty:
                    continue

                gen_subset = base_subset[
                    base_subset['Progeny_Group'].isin(sort_priority)
                ].copy()
                if not gen_subset.empty:
                    table_gen = gen_subset.pivot_table(
                        index='ID_Num', columns='Progeny_Group', values=metric)
                    table_gen = table_gen.reindex(
                        columns=[c for c in sort_priority if c in table_gen.columns])
                    table_gen.to_csv(
                        folder_genotype / f"{sex}_{age}Wks_{clean_metric}.csv")

                table_lin = base_subset.pivot_table(
                    index='ID_Num', columns='Progeny_Group', values=metric)
                sorted_cols = sorted(
                    table_lin.columns,
                    key=lambda x: next(
                        (idx for idx, g in enumerate(sort_priority) if x.startswith(g)), 99)
                )
                table_lin[sorted_cols].to_csv(
                    folder_lineage / f"{sex}_{age}Wks_{clean_metric}.csv")


def process_master_to_csv(master_path, csv_output_dir):
    print("Generating CSV analysis tables...")
    xl = pd.ExcelFile(master_path)
    all_data = []
    for sheet in xl.sheet_names:
        if sheet.lower() in ["summary", "notes", "calculations"]:
            continue
        raw_df = pd.read_excel(master_path, sheet_name=sheet)
        for start_idx in [0, 12]:
            subset = raw_df.iloc[:, start_idx:start_idx + 11].copy()
            subset.columns = ['Mouse Code'] + COLUMNS_TO_AVERAGE
            subset['Progeny_Group'] = sheet
            all_data.append(subset)

    final_df = pd.concat(all_data, ignore_index=True).dropna(
        subset=['Mouse Code'])
    parsed = final_df['Mouse Code'].apply(
        lambda x: pd.Series(parse_squashed_code(x)))
    final_df[['Gen_Label', 'Age_Extracted', 'Sex_Extracted', 'ID_Num']] = parsed
    generate_grouped_tables(final_df, csv_output_dir)


# --- 9. MAIN ---

if __name__ == "__main__":
    app = QApplication(sys.argv)

    dlg = ConfigDialog()
    if dlg.exec_() != QDialog.Accepted:
        sys.exit(0)

    config = dlg.get_config()
    study_name = config["study_name"]
    raw_data_root = Path(config["raw_data_root"])
    master_path = Path(config["master_file"])
    output_folder = Path(config["output_folder"])

    txt_files = list(raw_data_root.glob("*.txt"))
    analyzed_folder = raw_data_root / "Analyzed_Files"
    csv_output_dir = output_folder / f"{study_name}_CSVFiles"

    print(f"Study:      {study_name}")
    print(f"Raw data:   {raw_data_root}")
    print(f"Master:     {master_path}")
    print(f"CSV output: {csv_output_dir}")

    if not txt_files:
        QMessageBox.critical(None, "No Data Files",
                             f"No .txt files found in:\n{raw_data_root}")
        sys.exit(1)

    if not master_path.exists():
        QMessageBox.critical(None, "Master File Not Found",
                             f"Could not find:\n{master_path}")
        sys.exit(1)

    df_source = clean_and_average_data(txt_files)
    if df_source.empty:
        QMessageBox.critical(None, "No Valid Data",
                             "No valid data was found in the .txt files.")
        sys.exit(1)

    try:
        sync_to_master(df_source, master_path)

        analyzed_folder.mkdir(exist_ok=True)
        for file_path in txt_files:
            shutil.move(str(file_path), str(analyzed_folder / file_path.name))
        print(f"Processed files moved to {analyzed_folder}/")

        output_folder.mkdir(parents=True, exist_ok=True)
        process_master_to_csv(master_path, csv_output_dir)

        QMessageBox.information(
            None, "Pipeline Complete",
            f"All done!\n\n"
            f"CSV summaries saved to:\n{csv_output_dir}\n\n"
            f"Raw files archived to:\n{analyzed_folder}"
        )

    except PermissionError:
        QMessageBox.critical(None, "File In Use",
                             "Could not save the master file.\n"
                             "Please close the Excel file and try again.")
    except Exception as e:
        QMessageBox.critical(None, "Unexpected Error", str(e))
