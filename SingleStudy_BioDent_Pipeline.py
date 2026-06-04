import re
import pandas as pd
import openpyxl
import shutil
from pathlib import Path

from PyQt5.QtWidgets import QMessageBox

from BioDent_Utils import (
    COLUMNS_TO_AVERAGE, clean_and_average_data, ask_not_found_action
)

# Master header used by auto-created files
_MASTER_ID_HEADER = "Mouse Code"
_MASTER_HEADERS   = [_MASTER_ID_HEADER] + COLUMNS_TO_AVERAGE

_DEDUCE_SEX = 'Deduce from ID'


# --- 1. GROUP MAP HELPERS ---

def parse_id_and_group(mouse_id, group_map):
    """
    Return (prefix, group_name) for a mouse ID using flexible prefix matching
    (allows an optional separator between prefix and the numeric part, e.g.
    'ZC1M', 'ZC-1M', 'ZC_1M' all match prefix 'ZC').
    Returns (None, None) if no prefix matches.
    """
    if not mouse_id or (isinstance(mouse_id, float) and pd.isna(mouse_id)):
        return None, None
    code = str(mouse_id).strip()
    for prefix in sorted(group_map.keys(), key=len, reverse=True):
        if re.match(re.escape(prefix) + r'[-_.\s]?', code, re.IGNORECASE):
            val = group_map[prefix]
            group_name = val if isinstance(val, str) else val.get("group", prefix)
            return prefix, group_name
    return None, None


def _extract_sex_from_code(mouse_id):
    """Return 'Male', 'Female', or 'Unknown' by reading the trailing M/F in the ID."""
    code = str(mouse_id).strip().upper()
    m = re.search(r'[_\-\.]([MF])[_\-\.\d]', code) or re.search(r'([MF])\d*$', code)
    if m:
        return "Male" if m.group(1) == "M" else "Female"
    return "Unknown"


def deduce_sex(mouse_id, group_map):
    """
    Return 'Male', 'Female', or 'Unknown'.
    Priority: explicit sex in group_map entry → _DEDUCE_SEX sentinel (reads
    trailing M/F from the ID) → direct M/F scan of the full ID string.
    """
    prefix, _ = parse_id_and_group(mouse_id, group_map)
    if prefix:
        val = group_map.get(prefix, "")
        if isinstance(val, dict):
            s = val.get("sex", "")
            if s == _DEDUCE_SEX:
                return _extract_sex_from_code(mouse_id)
            if s in ("Male", "Female"):
                return s
    return _extract_sex_from_code(mouse_id)


# --- 2. MASTER SYNC ---

def _find_id_col(ws):
    """Return the 1-based column index of the Mouse Code column, defaulting to 1."""
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val and "mouse" in str(val).lower():
            return c
    return 1


def _find_metric_cols(ws):
    """
    Return a dict {column_name: 1-based col index} for each COLUMNS_TO_AVERAGE
    header found in row 1. Falls back to positional mapping if headers are absent.
    """
    header_map = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val and str(val).strip() in COLUMNS_TO_AVERAGE:
            header_map[str(val).strip()] = c

    if not header_map:
        # No recognized headers — assume cols 2…11 in order
        id_col = _find_id_col(ws)
        for j, col_name in enumerate(COLUMNS_TO_AVERAGE):
            header_map[col_name] = id_col + 1 + j

    return header_map


def sync_to_master(df_source, master_path):
    """
    Write averaged values into master_path.

    - If master does not exist: auto-create a flat single-sheet file with
      headers, populate IDs and values from df_source, and save.
    - If master exists but has no data rows yet (empty ID column): populate IDs
      from df_source, then write values.
    - If master has existing IDs: match by ID and write values. Unmatched IDs
      prompt the user to rename or skip (same as FKBP5 behaviour).
    """
    master_path = Path(master_path)

    # ---- Auto-create ----
    if not master_path.exists():
        print(f"Master not found — auto-creating: {master_path.name}")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        for j, h in enumerate(_MASTER_HEADERS, 1):
            ws.cell(row=1, column=j).value = h
        for i, (_, row_data) in enumerate(df_source.iterrows(), 2):
            ws.cell(row=i, column=1).value = str(row_data['ID'])
            for j, col_name in enumerate(COLUMNS_TO_AVERAGE, 2):
                ws.cell(row=i, column=j).value = row_data[col_name]
        wb.save(master_path)
        print(f"Master created with {len(df_source)} subjects.")
        return

    # ---- Pre-existing master ----
    wb = openpyxl.load_workbook(master_path)
    ws = wb.active
    id_col      = _find_id_col(ws)
    metric_cols = _find_metric_cols(ws)

    existing_ids = [
        str(ws.cell(row=r, column=id_col).value).strip()
        for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=id_col).value
    ]

    # If ID column is empty, populate IDs from df_source first
    if not existing_ids:
        for i, (_, row_data) in enumerate(df_source.iterrows(), 2):
            ws.cell(row=i, column=id_col).value = str(row_data['ID'])
        existing_ids = [str(row_data['ID']) for _, row_data in df_source.iterrows()]
        print(f"Populated {len(existing_ids)} IDs into empty master.")

    sync_count = 0
    synced_ids = set()
    all_processed_ids = set(df_source['ID'].astype(str).str.strip())

    for _, row_data in df_source.iterrows():
        mouse_code  = str(row_data['ID']).strip()
        found_match = False

        while True:
            for r in range(2, ws.max_row + 1):
                cell_val = ws.cell(row=r, column=id_col).value
                if cell_val and str(cell_val).strip().upper() == mouse_code.upper():
                    for col_name, col_idx in metric_cols.items():
                        ws.cell(row=r, column=col_idx).value = row_data.get(col_name)
                    sync_count += 1
                    synced_ids.add(mouse_code)
                    found_match = True
                    break

            if found_match:
                break

            new_name = ask_not_found_action(mouse_code)
            if new_name:
                mouse_code = new_name
            else:
                break

    wb.save(master_path)
    print(f"Synced {sync_count} subjects to master.")

    unmatched = all_processed_ids - synced_ids
    if unmatched:
        print("Subjects not found in master:")
        for m in sorted(unmatched):
            print(f"  - {m}")
    else:
        print("All subjects matched.")


# --- 3. CSV GENERATION ---

def process_master_to_csv(master_path, csv_output_dir, study_name, group_map, sex="", age=""):
    print("Generating CSV summaries...")
    master_path = Path(master_path)
    xl = pd.ExcelFile(master_path)
    all_data = []

    for sheet in xl.sheet_names:
        if sheet.lower() in ["summary", "notes", "calculations"]:
            continue
        raw_df = pd.read_excel(master_path, sheet_name=sheet, header=0)

        # Find ID column
        id_col_name = next(
            (c for c in raw_df.columns if "mouse" in str(c).lower()),
            raw_df.columns[0]
        )

        # Find metric columns — by header name first, then positional fallback
        metric_cols_found = [c for c in raw_df.columns if c in COLUMNS_TO_AVERAGE]
        if metric_cols_found:
            subset = raw_df[[id_col_name] + metric_cols_found].copy()
            subset.columns = ['Subject_ID'] + metric_cols_found
            # Pad any missing metrics with NaN
            for m in COLUMNS_TO_AVERAGE:
                if m not in subset.columns:
                    subset[m] = float('nan')
            subset = subset[['Subject_ID'] + COLUMNS_TO_AVERAGE]
        else:
            # Positional: col 0 = ID, cols 1–10 = metrics
            subset = raw_df.iloc[:, 0:11].copy()
            subset.columns = ['Subject_ID'] + COLUMNS_TO_AVERAGE

        subset['Sheet'] = sheet
        all_data.append(subset)

    if not all_data:
        print("No data sheets found in master.")
        return

    final_df = pd.concat(all_data, ignore_index=True).dropna(subset=['Subject_ID'])

    # Group assignment via flexible prefix matching
    if group_map:
        final_df['Group'] = final_df['Subject_ID'].apply(
            lambda x: (parse_id_and_group(str(x), group_map)[1] or "Unknown")
        )
    else:
        final_df['Group'] = final_df['Sheet']

    if sex == "Mixed" and group_map:
        final_df['Sex'] = final_df['Subject_ID'].apply(
            lambda x: deduce_sex(str(x), group_map)
        )

    csv_output_dir = Path(csv_output_dir)
    csv_output_dir.mkdir(parents=True, exist_ok=True)

    # Full summary CSV
    label_parts  = [study_name] + [p for p in [sex, age] if p and str(p).lower() not in ("none", "n/a", "")]
    summary_path = csv_output_dir / f"{'_'.join(label_parts)}_Summary.csv"
    summary_cols = (['Subject_ID', 'Group']
                    + (['Sex'] if 'Sex' in final_df.columns else [])
                    + COLUMNS_TO_AVERAGE)
    final_df[summary_cols].to_csv(summary_path, index=False)
    print(f"Summary saved to {summary_path}")

    # Per-metric pivot tables
    metrics_dir = csv_output_dir / "Per_Metric"
    metrics_dir.mkdir(exist_ok=True)

    if sex == "Mixed" and 'Sex' in final_df.columns:
        sex_splits = [("Male",   final_df[final_df['Sex'] == "Male"]),
                      ("Female", final_df[final_df['Sex'] == "Female"])]
    else:
        sex_splits = [(None, final_df)]

    for sex_label, df_split in sex_splits:
        out_dir = metrics_dir / sex_label if sex_label else metrics_dir
        out_dir.mkdir(exist_ok=True)
        if df_split.empty:
            print(f"No data for {sex_label}, skipping.")
            continue
        for metric in COLUMNS_TO_AVERAGE:
            clean_metric = re.sub(r'[\-\/\(\)]', '', metric).replace("  ", " ").strip()
            try:
                table = df_split.pivot_table(
                    index='Subject_ID', columns='Group', values=metric)
                table.to_csv(out_dir / f"{clean_metric}.csv")
            except Exception as e:
                print(f"Could not generate pivot for {metric}: {e}")

    print(f"Per-metric tables saved to {metrics_dir}")


# --- 4. PIPELINE ENTRY POINT ---

def run_pipeline(config):
    study_name    = config["study_name"]
    raw_data_root = Path(config["raw_data_root"])
    output_folder = Path(config["output_folder"])
    group_map     = config.get("group_map", {})
    sex           = config.get("sex", "")
    age           = config.get("age", "")

    # master_file is optional — blank/missing → auto-create in output_folder
    master_file_str = config.get("master_file", "").strip()
    if master_file_str:
        master_path = Path(master_file_str)
    else:
        safe_name   = re.sub(r'[^\w\-]', '_', study_name)
        master_path = output_folder / f"{safe_name}_Master.xlsx"

    txt_files       = list(raw_data_root.glob("*.txt"))
    analyzed_folder = raw_data_root / "Analyzed_Files"
    csv_output_dir  = output_folder / f"{study_name}_CSVFiles"

    print(f"Study:      {study_name}")
    print(f"Sex:        {sex or '(not set)'}")
    print(f"Age:        {age or '(not set)'}")
    print(f"Raw data:   {raw_data_root}")
    print(f"Master:     {master_path}")
    print(f"Group map:  {group_map}")
    print(f"CSV output: {csv_output_dir}")

    if not txt_files:
        if master_path.exists():
            reply = QMessageBox.question(
                None, "No New Data Files",
                f"No .txt files found in:\n{raw_data_root}\n\n"
                "Regenerate CSV summaries from the existing master file?",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                try:
                    csv_output_dir.mkdir(parents=True, exist_ok=True)
                    process_master_to_csv(
                        master_path, csv_output_dir, study_name, group_map, sex, age)
                    QMessageBox.information(
                        None, "Done",
                        f"CSV summaries saved to:\n{csv_output_dir}"
                    )
                except Exception as e:
                    QMessageBox.critical(None, "Unexpected Error", str(e))
        else:
            QMessageBox.critical(None, "No Data Files",
                                 f"No .txt files found in:\n{raw_data_root}")
        return

    df_source = clean_and_average_data(txt_files, group_map=group_map)
    if df_source.empty:
        QMessageBox.critical(
            None, "No Valid Data",
            "No valid data was found in the .txt files.\n"
            + ("Check that your group_map prefixes match the IDs in your files."
               if group_map else "Check that your .txt files contain valid data.")
        )
        return

    try:
        output_folder.mkdir(parents=True, exist_ok=True)
        sync_to_master(df_source, master_path)

        analyzed_folder.mkdir(exist_ok=True)
        for file_path in txt_files:
            shutil.move(str(file_path), str(analyzed_folder / file_path.name))
        print(f"Processed files moved to {analyzed_folder}/")

        process_master_to_csv(master_path, csv_output_dir, study_name, group_map, sex, age)

        QMessageBox.information(
            None, "Pipeline Complete",
            f"All done!\n\n"
            f"Master file:\n{master_path}\n\n"
            f"CSV summaries saved to:\n{csv_output_dir}\n\n"
            f"Raw files archived to:\n{analyzed_folder}"
        )

    except PermissionError:
        QMessageBox.critical(None, "File In Use",
                             "Could not save the master file.\n"
                             "Please close the Excel file and try again.")
    except Exception as e:
        QMessageBox.critical(None, "Unexpected Error", str(e))
