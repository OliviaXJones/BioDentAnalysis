import re
import os
import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import filedialog, ttk
from pathlib import Path
import shutil

# --- 1. SETTINGS & COLUMN MAPPING ---

# Use 'um' and 'uJ' consistently to match the cleaning logic
COLUMNS_TO_AVERAGE = [
    "1st Cycle Touchdown Distance (TDD 1st) - um",
    "Avg Loading Slope (Avg LS 1st-L) - N/um",
    "Avg Unloading Slope (Avg US 1st-L) - N/um",
    "Avg Energy Dissipated (Avg ED 3rd-L) - uJ",
    "Avg Creep Indentation Distance (Avg CID 1st-L) - um",
    "Indentation Distance Increase (IDI 1st-L) - um",
    "Total Indentation Distance (TID 1st-L) - um",
    "1st Cycle Creep Indentation Distance (CID 1st) - um",
    "1st Cycle Unloading Slope (US 1st) - N/um",
    "1st Cycle Indentation Distance (ID 1st) - um"
]


def get_config_paths():
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    print("Select your Master file (.xlsx or .xlsm)...")
    master_file = filedialog.askopenfilename(
        title="Select Master Excel File", filetypes=[("Excel Files", "*.xlsx *.xlsm")])
    print("Select the folder containing your .txt files...")
    txt_folder = filedialog.askdirectory(title="Select Folder with .txt Files")
    if not master_file or not txt_folder:
        print("Selection cancelled. Exiting.")
        exit()

    txt_folder_path = Path(txt_folder)
    analyzed_folder = txt_folder_path / "Analyzed_Files"
    txt_files = list(txt_folder_path.glob("*.txt"))
    return Path(master_file), txt_files, analyzed_folder


def get_manual_selection(mouse_id, group, target_col):
    selected_indices = []

    # Create a hidden root to handle the mainloop properly
    temp_root = tk.Tk()
    temp_root.withdraw()

    popup = tk.Toplevel(temp_root)
    popup.title(f"Data Selection: {mouse_id}")
    popup.attributes('-topmost', True)
    popup.geometry("1000x750")
    popup.lift()  # Force to front
    popup.focus_force()

    custom_font = ('Arial', 12)
    header_font = ('Arial', 12, 'bold')

    tk.Label(popup, text=f"Select the 4 rows to keep for {mouse_id}:",
             font=('Arial', 14, 'bold')).pack(pady=20)

    frame = tk.Frame(popup)
    frame.pack(padx=20, pady=10, fill="both", expand=True)
    vars_dict = {}

    headers = ["Idx", "Meas #", "ID 1st (um)", "Notes", "Selection"]
    widths = [8, 10, 15, 40, 15]
    for col, text in enumerate(headers):
        lbl = tk.Label(frame, text=text, font=header_font,
                       relief="groove", width=widths[col])
        lbl.grid(row=0, column=col, ipady=10)

    for i, (idx, row) in enumerate(group.iterrows()):
        grid_args = {'padx': 5, 'pady': 5, 'ipady': 10}
        tk.Label(frame, text=str(idx), font=custom_font).grid(
            row=i+1, column=0, **grid_args)
        tk.Label(frame, text=str(row['Measurement #']), font=custom_font).grid(
            row=i+1, column=1, **grid_args)
        tk.Label(frame, text=str(row[target_col]), font=custom_font).grid(
            row=i+1, column=2, **grid_args)
        tk.Label(frame, text=str(row['Notes'])[:40], font=custom_font, anchor="w").grid(
            row=i+1, column=3, **grid_args)

        var = tk.BooleanVar()
        ck = tk.Checkbutton(frame, variable=var, text="KEEP", font=header_font,
                            indicatoron=False, selectcolor="#90EE90", width=10)
        ck.grid(row=i+1, column=4, padx=10, pady=5, ipady=5)
        vars_dict[idx] = var

    def confirm_action():
        nonlocal selected_indices
        selected_indices = [idx for idx, v in vars_dict.items() if v.get()]
        popup.destroy()
        temp_root.destroy()

    btn = tk.Button(popup, text="Confirm Selection", font=('Arial', 13, 'bold'),
                    command=confirm_action, bg="#4CAF50", fg="white", padx=40, pady=15)
    btn.pack(pady=30)

    popup.grab_set()
    popup.wait_window()
    return selected_indices


def get_genotype_prefix(code):
    code_str = str(code).strip().upper()
    if 'H' in code_str:
        return "Heterozygous"
    elif 'W' in code_str:
        return "Wildtype"
    elif 'M' in code_str:
        return "Mutant"
    return None


def get_sex(code):
    return 'F' if 'F' in str(code).upper() else 'M'


def clean_and_average_data(txt_file_paths):
    all_data = []
    for path in txt_file_paths:
        try:
            temp_df = pd.read_csv(path, sep='\t', encoding='latin1')
            all_data.append(temp_df)
        except Exception as e:
            print(f"Skipping {path.name}: {e}")

    if not all_data:
        return pd.DataFrame()

    df = pd.concat(all_data, ignore_index=True)
    df.columns = [c.replace('µ', 'u') if isinstance(
        c, str) else c for c in df.columns]
    df = df.rename(columns={'Sample/Location': 'ID'})

    ID_REGEX = r"2?[WMH][.\-_]?\d+[.\-_]?[MF][.\-_]?\d+"

    def fix_mislabeled_id(row):
        notes = str(row['Notes']).strip().upper()
        match = re.search(ID_REGEX, notes)
        if match:
            true_id = re.sub(r"[.\-_]", "", match.group(0))
            if true_id != str(row['ID']).strip().upper():
                return true_id
        return row['ID']

    df['ID'] = df.apply(fix_mislabeled_id, axis=1)
    target_val_col = "1st Cycle Indentation Distance (ID 1st) - um"
    df = df.drop_duplicates(
        subset=['ID', 'Measurement #', target_val_col], keep='first')

    ignore_keywords = ["ignore", "do not use", "disregard", "don't", "ignor"]
    mask = df['Notes'].str.contains(
        '|'.join(ignore_keywords), case=False, na=False)
    df_cleaned = df[~mask].copy()

    df_cleaned = df_cleaned.dropna(subset=['ID'])
    df_cleaned['ID'] = df_cleaned['ID'].astype(str)

    final_averages = []
    unique_ids = sorted(df_cleaned['ID'].unique())
    i = 0
    while i < len(unique_ids):
        mouse_id = unique_ids[i]
        group = df_cleaned[df_cleaned['ID'] == mouse_id]
        mouse_id_str = str(mouse_id).strip()

        if len(group) > 4:
            indices = get_manual_selection(mouse_id_str, group, target_val_col)
            if indices:
                group = group.loc[indices]

        elif len(group) < 4 and len(group) > 0:
            print(f"\n🚨 {mouse_id_str} only has {len(group)} measurements.")
            action = input(
                "Action: [a]verage anyway, [s]kip, [r]ealign/rename: ").lower()
            if action == 'r':
                new_id = None
                for idx, row in group.iterrows():
                    scanned = fix_mislabeled_id(row)
                    if scanned != row['ID']:
                        new_id = scanned
                        break
                if not new_id:
                    new_id = input(
                        f"Enter the correct ID for {mouse_id_str}: ").strip().upper()
                if new_id:
                    new_id = re.sub(r"[.\-_]", "", new_id)
                    df_cleaned.loc[group.index, 'ID'] = new_id
                    unique_ids = sorted(df_cleaned['ID'].unique())
                    continue
            elif action == 's':
                i += 1
                continue

        for col in COLUMNS_TO_AVERAGE:
            group[col] = pd.to_numeric(group[col], errors='coerce')

        avg_values = group[COLUMNS_TO_AVERAGE].mean()
        avg_values['ID'] = mouse_id_str
        final_averages.append(avg_values)
        i += 1

    return pd.DataFrame(final_averages)


if __name__ == "__main__":
    MASTER_PATH, TXT_PATHS, ANALYZED_FOLDER = get_config_paths()
    df_source = clean_and_average_data(TXT_PATHS)

    if df_source.empty:
        print("No valid data found.")
        exit()

    try:
        wb = openpyxl.load_workbook(MASTER_PATH)
        sync_count = 0

# --- THE SYNCING LOOP ---
        synced_ids = set()
        all_processed_ids = set(df_source['ID'].astype(str).str.strip())

        for _, row_data in df_source.iterrows():
            mouse_code = str(row_data['ID']).strip()

            # This inner loop allows for one "retry" if the name is wrong
            while True:
                genotype = get_genotype_prefix(mouse_code)
                sex = get_sex(mouse_code)

                if not genotype:
                    print(f"❌ Could not determine genotype for {mouse_code}")
                    break

                target_sheets = [s for s in wb.sheetnames if s.strip(
                ).lower().startswith(genotype.lower())]
                found_match = False

                for sheet_name in target_sheets:
                    ws = wb[sheet_name]
                    start_col = 1 if sex == 'M' else 13

                    for r in range(2, ws.max_row + 1):
                        cell_val = ws.cell(row=r, column=start_col).value
                        if cell_val and str(cell_val).strip().lower() == mouse_code.lower():

                            # SKIP CHECK
                            existing_entry = ws.cell(
                                row=r, column=start_col + 1).value
                            if existing_entry is not None:

                                synced_ids.add(mouse_code)
                                found_match = True
                                break

                            for i, col_name in enumerate(COLUMNS_TO_AVERAGE):
                                ws.cell(row=r, column=start_col + 1 +
                                        i).value = row_data[col_name]

                            print(f"✅ [SYNCED] {mouse_code}")
                            sync_count += 1
                            synced_ids.add(mouse_code)
                            found_match = True
                            break
                    if found_match:
                        break

                if found_match:
                    break  # Success, move to next mouse in df_source
                else:
                    print(
                        f"\n🔍 [NOT FOUND] '{mouse_code}' was not found in the Master Excel.")
                    choice = input(
                        f"Action for {mouse_code}: [r]ename/retry, [s]kip: ").lower()
                    if choice == 'r':
                        mouse_code = input(
                            f"Enter the ID EXACTLY as it appears in Excel: ").strip()
                        # Loop repeats once more with the new mouse_code
                    else:
                        break  # Move to next mouse

        # --- REPORT MISSING MICE ---
        unmatched_ids = all_processed_ids - synced_ids

        if unmatched_ids:
            print(
                "\n⚠️  THE FOLLOWING MICE WERE PROCESSED BUT NOT FOUND IN THE MASTER LIST:")
            for m_id in sorted(unmatched_ids):
                print(f"   - {m_id}")
        else:
            print("\n✨ All processed mice were successfully matched in the Master list.")

        wb.save(MASTER_PATH)
        print(f"\n✅ Done! Successfully synced {sync_count} mice.")

        ANALYZED_FOLDER.mkdir(exist_ok=True)
        for file_path in TXT_PATHS:
            shutil.move(str(file_path), str(ANALYZED_FOLDER / file_path.name))
        print(f"📂 Processed files moved to {ANALYZED_FOLDER.name}.")

    except PermissionError:
        print("\n❌ Error: Close the Excel file and try again.")
    except Exception as e:
        print(f"\n❌ An error occurred: {e}")
