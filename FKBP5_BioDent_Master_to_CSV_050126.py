import re
import os
import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import filedialog, ttk
from pathlib import Path
import shutil

# --- 1. SETTINGS & COLUMN MAPPING ---

COLUMNS_TO_AVERAGE = [
    "1st Cycle Touchdown Distance (TDD 1st) - um",
    "Avg Loading Slope (Avg LS 1st-L) - N/um",
    "Avg Unloading Slope (Avg US 1st-L) - N/um",
    "Avg Energy Dissipated (Avg ED 3rd-L) - uJ",
    "Avg Creep Indentation Distance (Avg CID 1st-L) - um",
    "Indentation Distance Increase (IDI 1st-L) - um",
    "Total Indentation Distance (TID 1st-L) - um",
    "1st Cycle Creep Indentation Distance (CID 1st) - um",
    "1st Cycle Unloading Slope (US 1st) - N/um",
    "1st Cycle Indentation Distance (ID 1st) - um"
]

# --- 2. HELPER FUNCTIONS ---


def parse_squashed_code(code):
    """Parses IDs like 'W12M10', 'H24F5', or '2W10M10' for CSV grouping"""
    code = str(code).strip().upper()
    match = re.match(r"(\d+)?([A-Z])(\d+)([MF])(\d+)", code)
    if match:
        gen_num, gen_char, age, sex_char, mouse_num = match.groups()
        gen_map = {'W': 'Wildtype', 'M': 'Mutant', 'H': 'Heterozygous'}
        genotype = gen_map.get(gen_char, "Unknown")
        sex = "Male" if sex_char == 'M' else "Female"
        if gen_num:
            genotype = f"{gen_num}{genotype}"
        return genotype, age, sex, mouse_num
    return None, None, None, None


def get_config_paths():
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    print("Select your Master file (.xlsx or .xlsm)...")
    master_file = filedialog.askopenfilename(
        title="Select Master Excel File", filetypes=[("Excel Files", "*.xlsx *.xlsm")])
    print("Select the folder containing your .txt files...")
    txt_folder = filedialog.askdirectory(title="Select Folder with .txt Files")
    if not master_file or not txt_folder:
        print("Selection cancelled. Exiting.")
        exit()

    txt_folder_path = Path(txt_folder)
    analyzed_folder = txt_folder_path / "Analyzed_Files"
    txt_files = list(txt_folder_path.glob("*.txt"))
    return Path(master_file), txt_files, analyzed_folder


def get_manual_selection(mouse_id, group, target_col):
    selected_indices = []
    temp_root = tk.Tk()
    temp_root.withdraw()
    popup = tk.Toplevel(temp_root)
    popup.title(f"Data Selection: {mouse_id}")
    popup.attributes('-topmost', True)
    popup.geometry("1000x750")

    tk.Label(popup, text=f"Select the 4 rows to keep for {mouse_id}:", font=(
        'Arial', 14, 'bold')).pack(pady=20)
    frame = tk.Frame(popup)
    frame.pack(padx=20, pady=10, fill="both", expand=True)
    vars_dict = {}

    headers = ["Idx", "Meas #", "ID 1st (um)", "Notes", "Selection"]
    for col, text in enumerate(headers):
        tk.Label(frame, text=text, font=('Arial', 12, 'bold'),
                 relief="groove", width=15).grid(row=0, column=col)

    for i, (idx, row) in enumerate(group.iterrows()):
        tk.Label(frame, text=str(idx)).grid(row=i+1, column=0)
        tk.Label(frame, text=str(row['Measurement #'])).grid(row=i+1, column=1)
        tk.Label(frame, text=str(row[target_col])).grid(row=i+1, column=2)
        tk.Label(frame, text=str(row['Notes'])[:40]).grid(row=i+1, column=3)
        var = tk.BooleanVar()
        tk.Checkbutton(frame, variable=var, text="KEEP", indicatoron=False,
                       selectcolor="#90EE90", width=10).grid(row=i+1, column=4)
        vars_dict[idx] = var

    def confirm_action():
        nonlocal selected_indices
        selected_indices = [idx for idx, v in vars_dict.items() if v.get()]
        popup.destroy()
        temp_root.destroy()

    tk.Button(popup, text="Confirm Selection", command=confirm_action,
              bg="#4CAF50", fg="white", padx=40, pady=15).pack(pady=30)
    popup.grab_set()
    popup.wait_window()
    return selected_indices


def get_genotype_prefix(code):
    if not code:
        return ""
    genotype_full, age, sex, mouse_num = parse_squashed_code(code)

    if genotype_full:
        # Returns "Wildtype", "Mutant", etc.
        return re.sub(r'\d+', '', genotype_full)
    return ""  # Return empty string instead of None


def get_sex(code):
    return 'F' if 'F' in str(code).upper() else 'M'

# --- 3. DATA CLEANING & AVERAGING ---


def clean_and_average_data(txt_file_paths):
    all_data = []
    for path in txt_file_paths:
        try:
            temp_df = pd.read_csv(path, sep='\t', encoding='latin1')
            all_data.append(temp_df)
        except Exception as e:
            print(f"Skipping {path.name}: {e}")

    if not all_data:
        return pd.DataFrame()

    df = pd.concat(all_data, ignore_index=True)
    df.columns = [c.replace('µ', 'u') if isinstance(
        c, str) else c for c in df.columns]
    df = df.rename(columns={'Sample/Location': 'ID'})

    ID_REGEX = r"2?[WMH][.\-_]?\d+[.\-_]?[MF][.\-_]?\d+"

    def fix_mislabeled_id(row):
        notes = str(row['Notes']).strip().upper()
        match = re.search(ID_REGEX, notes)
        if match:
            true_id = re.sub(r"[.\-_]", "", match.group(0))
            if true_id != str(row['ID']).strip().upper():
                return true_id
        return row['ID']

    df['ID'] = df.apply(fix_mislabeled_id, axis=1)
    target_val_col = "1st Cycle Indentation Distance (ID 1st) - um"
    df = df.drop_duplicates(
        subset=['ID', 'Measurement #', target_val_col], keep='first')

    ignore_keywords = ["ignore", "do not use", "disregard", "don't", "ignor"]
    mask = df['Notes'].str.contains(
        '|'.join(ignore_keywords), case=False, na=False)
    df_cleaned = df[~mask].copy()
    df_cleaned = df_cleaned.dropna(subset=['ID'])
    df_cleaned['ID'] = df_cleaned['ID'].astype(str)

    final_averages = []
    unique_ids = sorted(df_cleaned['ID'].unique())
    i = 0
    while i < len(unique_ids):
        mouse_id = unique_ids[i]
        group = df_cleaned[df_cleaned['ID'] == mouse_id]
        mouse_id_str = str(mouse_id).strip()

        if len(group) > 4:
            indices = get_manual_selection(mouse_id_str, group, target_val_col)
            if indices:
                group = group.loc[indices]

        elif len(group) < 4 and len(group) > 0:
            print(f"\n🚨 {mouse_id_str} only has {len(group)} measurements.")
            action = input(
                "Action: [a]verage anyway, [s]kip, [r]ealign/rename: ").lower()
            if action == 'r':
                new_id = input(
                    f"Enter the correct ID for {mouse_id_str}: ").strip().upper()
                if new_id:
                    new_id = re.sub(r"[.\-_]", "", new_id)
                    df_cleaned.loc[group.index, 'ID'] = new_id
                    unique_ids = sorted(df_cleaned['ID'].unique())
                    continue
            elif action == 's':
                i += 1
                continue

        for col in COLUMNS_TO_AVERAGE:
            group[col] = pd.to_numeric(group[col], errors='coerce')

        avg_values = group[COLUMNS_TO_AVERAGE].mean()
        avg_values['ID'] = mouse_id_str
        final_averages.append(avg_values)
        i += 1

    return pd.DataFrame(final_averages)

# --- 4. CSV EXPORT LOGIC ---


def generate_grouped_tables(df, base_dir):
    sort_priority = ['Wildtype', 'Mutant', 'Heterozygous']
    folder_genotype = base_dir / "Analysis_By_Genotype"
    folder_lineage = base_dir / "Analysis_By_Lineage"
    folder_genotype.mkdir(parents=True, exist_ok=True)
    folder_lineage.mkdir(parents=True, exist_ok=True)

    for metric in COLUMNS_TO_AVERAGE:
        clean_metric = re.sub(
            r'[\-\/\(\)]', '', metric).replace("  ", " ").strip()
        for age in df['Age_Extracted'].dropna().unique():
            for sex in ['Male', 'Female']:
                base_subset = df[(df['Sex_Extracted'] == sex) & (
                    df['Age_Extracted'] == age)].copy()
                if base_subset.empty:
                    continue

                # Set 1: Exact Genotypes
                gen_subset = base_subset[base_subset['Progeny_Group'].isin(
                    sort_priority)].copy()
                if not gen_subset.empty:
                    table_gen = gen_subset.pivot_table(
                        index='ID_Num', columns='Progeny_Group', values=metric)
                    table_gen = table_gen.reindex(
                        columns=[c for c in sort_priority if c in table_gen.columns])
                    table_gen.to_csv(folder_genotype /
                                     f"{sex}_{age}Wks_{clean_metric}.csv")

                # Set 2: All Lineages
                table_lin = base_subset.pivot_table(
                    index='ID_Num', columns='Progeny_Group', values=metric)
                sorted_cols = sorted(table_lin.columns, key=lambda x: next(
                    (i for i, g in enumerate(sort_priority) if x.startswith(g)), 99))
                table_lin[sorted_cols].to_csv(
                    folder_lineage / f"{sex}_{age}Wks_{clean_metric}.csv")


def process_master_to_csv(master_path):
    print("\n📊 Generating CSV analysis tables...")
    xl = pd.ExcelFile(master_path)
    all_data = []
    for sheet in xl.sheet_names:
        if sheet.lower() in ["summary", "notes", "calculations"]:
            continue
        raw_df = pd.read_excel(master_path, sheet_name=sheet)
        for start_idx in [0, 12]:
            subset = raw_df.iloc[:, start_idx:start_idx+11].copy()
            subset.columns = ['Mouse Code'] + COLUMNS_TO_AVERAGE
            subset['Progeny_Group'] = sheet
            all_data.append(subset)

    final_df = pd.concat(all_data, ignore_index=True).dropna(
        subset=['Mouse Code'])
    parsed = final_df['Mouse Code'].apply(
        lambda x: pd.Series(parse_squashed_code(x)))
    final_df[['Gen_Label', 'Age_Extracted', 'Sex_Extracted', 'ID_Num']] = parsed
    generate_grouped_tables(final_df, master_path.parent)

# --- 5. MAIN WORKFLOW ---


if __name__ == "__main__":
    MASTER_PATH, TXT_PATHS, ANALYZED_FOLDER = get_config_paths()
    df_source = clean_and_average_data(TXT_PATHS)

    if df_source.empty:
        print("No valid data found.")
        exit()

    try:
        wb = openpyxl.load_workbook(MASTER_PATH)
        sync_count = 0
        synced_ids = set()
        all_processed_ids = set(df_source['ID'].astype(str).str.strip())

        # Get all valid data sheets (skipping summary/notes)
        all_sheets = [s for s in wb.sheetnames if s.lower(
        ) not in ["summary", "notes", "calculations"]]

        for _, row_data in df_source.iterrows():
            mouse_code = str(row_data['ID']).strip()
            found_match = False

            while True:
                genotype = get_genotype_prefix(mouse_code)
                sex = get_sex(mouse_code)
                start_col = 1 if sex == 'M' else 13

                # 1. Prioritize sheets matching the genotype
                target_sheets = [
                    s for s in all_sheets if genotype.lower() in s.lower()]

                # 2. Add all other sheets as backup (in case a Mutant is in a Het sheet)
                other_sheets = [
                    s for s in all_sheets if s not in target_sheets]
                search_order = target_sheets + other_sheets

                for sheet_name in search_order:
                    ws = wb[sheet_name]
                    # Search up to row 100 (or ws.max_row)
                    for r in range(2, max(ws.max_row + 1, 50)):
                        cell_val = ws.cell(row=r, column=start_col).value

                        # Clean both values for a fair comparison
                        clean_cell = str(cell_val).strip(
                        ).upper() if cell_val else ""
                        clean_target = mouse_code.strip().upper()

                        if clean_cell == clean_target:
                            # Write the data
                            for i, col_name in enumerate(COLUMNS_TO_AVERAGE):
                                ws.cell(row=r, column=start_col + 1 +
                                        i).value = row_data[col_name]

                            print(
                                f"✅ [MATCHED] {mouse_code} in {sheet_name} (Row {r})")
                            sync_count += 1
                            synced_ids.add(mouse_code)
                            found_match = True
                            break
                    if found_match:
                        break

                if found_match:
                    break
                else:
                    print(
                        f"\n🔍 [NOT FOUND] '{mouse_code}' not found in any sheet.")
                    choice = input(
                        f"Action for {mouse_code}: [r]ename/retry, [s]kip: ").lower()
                    if choice == 'r':
                        mouse_code = input(
                            f"Enter the ID EXACTLY as it appears in Excel: ").strip()
                    else:
                        break

        wb.save(MASTER_PATH)
        print(f"\n✅ Done! Successfully synced {sync_count} mice.")
        # Reporting
        unmatched_ids = all_processed_ids - synced_ids
        if unmatched_ids:
            print(
                "\n⚠️  THE FOLLOWING MICE WERE PROCESSED BUT NOT FOUND IN THE MASTER LIST:")
            for m_id in sorted(unmatched_ids):
                print(f"   - {m_id}")
        else:
            print("\n✨ All processed mice were successfully matched in the Master list.")

        wb.save(MASTER_PATH)
        print(f"\n✅ Done! Successfully synced {sync_count} mice.")

        # Cleanup and CSV Generation
        ANALYZED_FOLDER.mkdir(exist_ok=True)
        for file_path in TXT_PATHS:
            shutil.move(str(file_path), str(ANALYZED_FOLDER / file_path.name))

        process_master_to_csv(MASTER_PATH)
        print("📂 Processed files moved and CSV summaries generated.")

    except PermissionError:
        print("\n❌ Error: Close the Excel file and try again.")
    except Exception as e:
        print(f"\n❌ An error occurred: {e}")
